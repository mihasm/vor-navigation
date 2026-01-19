from openpyxl import load_workbook

import os

this_dir, this_filename = os.path.split(__file__)
DATA_PATH = os.path.join(this_dir, "data")

class Airport:
	def __init__(self,icao,iata,rwy_length,lat,lng,\
				 mag_var,elevation,name,city,state,country,\
				 prec_app,timezone,region):

		self.icao = icao
		self.iata = iata
		self.rwy_length = rwy_length
		self.lat = lat
		self.lng = lng
		self.mag_var = mag_var
		self.elevation = elevation
		self.name = name
		self.city = city
		self.state = state
		self.country = country
		self.prec_app = prec_app
		self.timezone = timezone
		self.region = region

def get_airports():
	airports = []
	wb = load_workbook(filename = os.path.join(DATA_PATH,"airports.xlsx"))
	ws = wb["Airports"]
	list_airports = [[cell.value for cell in row] for row in ws.iter_rows()]

	for i in range(1,len(list_airports)):
		icao,iata,rwy_length,lat,lng,mag_var,\
		elevation,name,city,state,country,_,_,_,_,_,\
		prec_app,_,_,_,timezone,region,_,_,_,_ = list_airports[i]

		lat = lat.replace("N","+").replace("S","-")
		#lat = float(lat)/1e6
		lng = lng.replace("E","+").replace("W","-")
		#lng = float(lng)/1e6
		sign_lat,lat_degs,lat_mins,lat_secs,lat_hundr = lat[0],float(lat[1:3]),float(lat[3:5]),float(lat[5:7]),float(lat[7:9])
		sign_lng,lng_degs,lng_mins,lng_secs,lng_hundr = lng[0],float(lng[1:4]),float(lng[4:6]),float(lng[6:8]),float(lng[8:10])

		lat = convert_position(sign_lat,lat_degs,lat_mins,lat_secs,lat_hundr)
		lng = convert_position(sign_lng,lng_degs,lng_mins,lng_secs,lng_hundr)

		a = Airport(icao,iata,rwy_length,lat,lng,mag_var,elevation,name,city,state,country,prec_app,timezone,region)
		airports.append(a)
	return airports

def convert_position(sign,degs,mins,secs,hundr):
	out = degs+mins/60+secs/60/100+hundr/10000
	if sign == "-":
		out = out*-1
	return out